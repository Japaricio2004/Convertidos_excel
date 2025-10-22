from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import io
import os
import traceback
import re
import numpy as np
from openpyxl import load_workbook
from datetime import datetime

app = Flask(__name__)

# Memoria simple para almacenar el último libro cargado como dict de DataFrames
_LAST_BOOK = {}


def _clean_column_name(name: str) -> str:
    """
    Limpia etiquetas de encabezado eliminando cualquier rastro de 'Unnamed: ...',
    normaliza espacios y quita separadores redundantes.
    """
    if name is None:
        return ""
    s = str(name)
    # Eliminar patrones Unnamed: <algo> en cualquier nivel
    # Casos comunes: 'Unnamed: 0_level_1', 'Unnamed: 2', etc.
    import re
    s = re.sub(r"Unnamed:\s*[^-_|,;]*", "", s, flags=re.IGNORECASE)
    # Eliminar duplicidad de separadores cuando quedan como ' - ' al principio/fin
    s = s.replace("  ", " ").strip()
    # Remover separadores sueltos al inicio/fin
    s = re.sub(r"^(?:-|–|—|·|\||:)+\s*", "", s)
    s = re.sub(r"\s*(?:-|–|—|·|\||:)\s*$", "", s)
    # Colapsar múltiples espacios y guiones
    s = re.sub(r"\s{2,}", " ", s)
    s = re.sub(r"\s*-\s*", " - ", s)
    return s.strip()


def _clean_columns(cols):
    return [_clean_column_name(c) for c in cols]

@app.route("/")
def index():
    return render_template("index.html")


def _choose_engine(filename: str) -> str:
    """Selecciona el engine correcto según la extensión del archivo."""
    fname = filename.lower()
    if fname.endswith((".xlsx", ".xlsm")):
        return "openpyxl"
    if fname.endswith(".xls"):
        return "xlrd"
    if fname.endswith(".xlsb"):
        return "pyxlsb"
    # Por defecto, intentar openpyxl
    return "openpyxl"


def _promote_header(df: pd.DataFrame) -> pd.DataFrame:
    """
    Promueve la fila más informativa como encabezado para evitar columnas 'Unnamed'.
    Estrategia:
    - Encontrar la fila con mayor cantidad de valores no vacíos.
    - Usarla como encabezado, limpiar espacios, reemplazar NaN por cadena vacía.
    - Remover filas hasta ese encabezado y resetear índice.
    """
    if df.empty:
        return df

    # Reemplazar todo-NaN por vacío para contar mejor
    counts = df.notna().sum(axis=1)
    header_row_idx = int(counts.idxmax())

    # Si la primera fila ya parece ser la mejor, igual normalizamos
    new_header = df.iloc[header_row_idx].fillna("")
    # Convertir a string y limpiar
    new_columns = [str(c).strip() for c in new_header]

    # Si todos están vacíos, no cambiamos nada
    if all(col == "" for col in new_columns):
        return df.fillna("")

    # Crear nuevo dataframe sin la fila de encabezado
    body = df.iloc[header_row_idx + 1 :].reset_index(drop=True)
    body.columns = new_columns

    # Limpiar NaN
    body = body.fillna("")

    return body


@app.route("/formula", methods=["POST"])
def formula():
    try:
        data = request.get_json(force=True)
        sheet = data.get("sheet")
        expr = data.get("expr", "").strip()
        name = data.get("name", "Resultado")
        mode = data.get("mode", "row")  # 'row' o 'total'
        fmt = data.get("format")  # e.g., 'percent:0.00', 'number:0.00', 'auto'

        if not _LAST_BOOK:
            return jsonify({"error": "Primero sube un archivo Excel."}), 400
        if not sheet or sheet not in _LAST_BOOK:
            return jsonify({"error": "Hoja no encontrada."}), 400
        if not expr:
            return jsonify({"error": "Expresión vacía."}), 400

        df = _LAST_BOOK[sheet].copy()

        # Mapeo de funciones estilo Excel a numpy/pandas (robusto a tipos mixtos)
        def _to_series_list(args):
            out = []
            for a in args:
                if isinstance(a, pd.Series):
                    out.append(pd.to_numeric(a, errors='coerce'))
                elif isinstance(a, (list, tuple, np.ndarray)):
                    out.append(pd.to_numeric(pd.Series(a), errors='coerce'))
                else:
                    # Escalar a serie constante
                    try:
                        val = pd.to_numeric(a, errors='coerce')
                    except Exception:
                        val = np.nan
                    out.append(pd.Series([val] * len(df)))
            return out

        def _row_df(args):
            return pd.concat(_to_series_list(args), axis=1)

        func_map = {
            'SUM': lambda *args: (_row_df(args).sum(axis=1, skipna=True) if mode=='row' else float(np.nansum([pd.to_numeric(a, errors='coerce').sum(skipna=True) if isinstance(a, pd.Series) else pd.to_numeric(a, errors='coerce') for a in args]))),
            'AVERAGE': lambda *args: (_row_df(args).mean(axis=1, skipna=True) if mode=='row' else (
                (sum([pd.to_numeric(a, errors='coerce').sum(skipna=True) if isinstance(a, pd.Series) else (pd.to_numeric(a, errors='coerce') if pd.notna(pd.to_numeric(a, errors='coerce')) else 0) for a in args])) /
                max(1, sum([ (pd.to_numeric(a, errors='coerce').notna().sum() if isinstance(a, pd.Series) else (1 if pd.notna(pd.to_numeric(a, errors='coerce')) else 0)) for a in args ]))
            )),
            'MIN': lambda *args: (_row_df(args).min(axis=1, skipna=True) if mode=='row' else (
                np.nanmin(pd.concat(_to_series_list(args), axis=0).to_numpy())
            )),
            'MAX': lambda *args: (_row_df(args).max(axis=1, skipna=True) if mode=='row' else (
                np.nanmax(pd.concat(_to_series_list(args), axis=0).to_numpy())
            )),
            'COUNT': lambda *args: (_row_df(args).notna().sum(axis=1) if mode=='row' else sum([
                (pd.to_numeric(a, errors='coerce').notna().sum() if isinstance(a, pd.Series) else (1 if pd.notna(pd.to_numeric(a, errors='coerce')) else 0)) for a in args
            ])),
        }

        # Funciones en español estilo Excel
        def _ensure_series(x):
            if isinstance(x, pd.Series):
                return x
            try:
                return pd.Series([x]*len(df))
            except Exception:
                return pd.Series([np.nan]*len(df))

        def __SI__(cond, v_true, v_false):
            cond_s = _ensure_series(cond).astype(bool)
            a = _ensure_series(v_true)
            b = _ensure_series(v_false)
            return pd.Series(np.where(cond_s, a, b)) if mode=='row' else (v_true if bool(cond) else v_false)

        def __CONCATENAR__(*args):
            cols = [_ensure_series(a).astype(str).fillna('') for a in args]
            if mode=='row':
                out = cols[0]
                for c in cols[1:]:
                    out = out + c
                return out
            else:
                return ''.join(str(a) for a in args)

        def __REDONDEAR__(num, dec):
            if mode=='row':
                n = int(float(dec)) if not isinstance(dec, pd.Series) else int(float(dec.iloc[0] or 0))
                return pd.to_numeric(_ensure_series(num), errors='coerce').round(n)
            else:
                return round(float(num), int(float(dec)))

        def __MAYUSC__(txt):
            if mode=='row':
                return _ensure_series(txt).astype(str).str.upper()
            return str(txt).upper()

        def __NOMPROPIO__(txt):
            if mode=='row':
                return _ensure_series(txt).astype(str).str.title()
            return str(txt).title()

        def __LARGO__(txt):
            if mode=='row':
                return _ensure_series(txt).astype(str).str.len()
            return len(str(txt))

        def __ALEATORIO_ENTRE__(a, b):
            try:
                lo = int(float(a if not isinstance(a, pd.Series) else a.fillna(0).iloc[0]))
                hi = int(float(b if not isinstance(b, pd.Series) else b.fillna(0).iloc[0]))
            except Exception:
                lo, hi = 0, 1
            if mode=='row':
                size = len(df)
                return pd.Series(np.random.randint(lo, hi+1, size=size))
            else:
                return int(np.random.randint(lo, hi+1))

        def __DIAS_LAB__(f1, f2, feriados=None):
            feriados = feriados or []
            def to_np_date(x):
                x = pd.to_datetime(x, errors='coerce')
                return np.datetime64(x.date()) if pd.notna(x) else None
            def count_days(a, b):
                a = pd.to_datetime(a, errors='coerce')
                b = pd.to_datetime(b, errors='coerce')
                if pd.isna(a) or pd.isna(b):
                    return np.nan
                hol = [np.datetime64(pd.to_datetime(h, errors='coerce').date()) for h in (feriados if isinstance(feriados, (list, tuple)) else []) if pd.notna(pd.to_datetime(h, errors='coerce'))]
                try:
                    return np.busday_count(a.date(), b.date(), weekmask='1111100', holidays=hol)
                except Exception:
                    return np.nan
            if mode=='row':
                s1 = _ensure_series(f1)
                s2 = _ensure_series(f2)
                return pd.Series([count_days(a,b) for a,b in zip(s1, s2)])
            else:
                return float(count_days(f1, f2))

        def __BUSCARV__(lookup, *args):
            # args puede ser (col1, col2, ..., idx[, exact]) o (cols_tuple, idx[, exact])
            exact = True
            cols = []
            # Extraer idx y exact si vienen al final
            args_list = list(args)
            if len(args_list) >= 2 and isinstance(args_list[-1], (bool, np.bool_)):
                exact = bool(args_list.pop())
            # idx
            if not args_list:
                return pd.Series([np.nan]*len(df)) if mode=='row' else np.nan
            idx = args_list.pop()
            try:
                idx = int(float(idx))
            except Exception:
                idx = 1
            # Rango de columnas
            if len(args_list)==1 and isinstance(args_list[0], pd.Series) is False:
                # puede venir como string de nombres, pero con el parser actual no empaquetamos tuplas
                pass
            # Construir lista de series de columnas desde args_list
            # Aceptar que vengan como referencias de columna (Series) en orden
            for a in args_list:
                if isinstance(a, pd.Series):
                    cols.append(a)
            if len(cols) < 2:
                return pd.Series([np.nan]*len(df)) if mode=='row' else np.nan
            idx = max(1, min(idx, len(cols)))
            key_series = pd.to_numeric(cols[0], errors='ignore')
            ret_series = cols[idx-1]
            # Construir mapping
            map_df = pd.DataFrame({f'c{i}': c for i, c in enumerate(cols)})
            mapping = pd.Series(map_df[f'c{idx-1}'].values, index=map_df['c0']).to_dict()
            lk = lookup if isinstance(lookup, pd.Series) else _ensure_series(lookup)
            res = lk.map(mapping)
            if not exact:
                # para aproximado, devolver exacto si existe, else NaN (implementación simple)
                return res
            return res

        # Añadir alias en español que apunten a implementaciones
        func_map_es = {
            '__PROMEDIO__': func_map['AVERAGE'],
            '__SI__': __SI__,
            '__CONCATENAR__': __CONCATENAR__,
            '__DIAS_LAB__': __DIAS_LAB__,
            '__REDONDEAR__': __REDONDEAR__,
            '__ALEATORIO_ENTRE__': __ALEATORIO_ENTRE__,
            '__MAYUSC__': __MAYUSC__,
            '__NOMPROPIO__': __NOMPROPIO__,
            '__LARGO__': __LARGO__,
            '__BUSCARV__': __BUSCARV__,
        }

        # Reemplazar nombres de columnas por variables seguras
        local_ctx = {}
        for col in df.columns:
            var = re.sub(r"[^A-Za-z0-9_]", "_", col)
            # asegurar único
            base = var
            k = 1
            while var in local_ctx:
                var = f"{base}_{k}"
                k += 1
            local_ctx[var] = pd.to_numeric(df[col], errors='coerce')
            # Permitir también acceso por nombre con comillas: [Col Name]
        
        # Helpers
        def col_ref(token: str):
            t = token.strip()
            if t in local_ctx:
                return local_ctx[t]
            # Intentar como texto entre [] o ''
            t2 = t.strip("[]'\"")
            key = re.sub(r"[^A-Za-z0-9_]", "_", t2)
            series = local_ctx.get(key)
            if series is None:
                # Devolver serie de NaN para evitar errores tipo NoneType en operaciones
                series = pd.Series([np.nan] * len(df))
            return series

        # Preprocesar expresión: permitir funciones Excel y nombres de columnas sin prefijo
        expr_py = expr
        # Normalizar separadores decimales y %
        expr_py = expr_py.replace(";", ",")
        # Normalizar funciones en español con y sin acento y con punto
        replacements = [
            (r"\bPROMEDIO\s*\(", "__PROMEDIO__("),
            (r"\bSI\s*\(", "__SI__("),
            (r"\bCONCATENAR\s*\(", "__CONCATENAR__("),
            (r"D[ÍI]AS[\._]LAB\s*\(", "__DIAS_LAB__("),
            (r"\bREDONDEAR\s*\(", "__REDONDEAR__("),
            (r"ALEATORIO[\._]ENTRE\s*\(", "__ALEATORIO_ENTRE__("),
            (r"\bMAYUSC\s*\(", "__MAYUSC__("),
            (r"\bNOMPROPIO\s*\(", "__NOMPROPIO__("),
            (r"\bLARGO\s*\(", "__LARGO__("),
            (r"\bBUSCARV\s*\(", "__BUSCARV__("),
        ]
        for pat, rep in replacements:
            expr_py = re.sub(pat, rep, expr_py, flags=re.IGNORECASE)
        # Reemplazo de funciones en inglés existentes
        for fn in func_map.keys():
            expr_py = re.sub(rf"\b{fn}\s*\(", f"__{fn}__(", expr_py, flags=re.IGNORECASE)
        
        # Tokenizer simple para identificar identificadores y mapear a variables del contexto
        tokens = re.findall(r"[A-Za-z_][A-Za-z0-9_]*|\d+(?:\.\d+)?|[()+\-*/%,]|\[.*?\]", expr_py)
        built = []
        for tok in tokens:
            up = tok.upper()
            if up in {f"__{k}__" for k in func_map.keys()}:
                built.append(tok.upper())
            elif re.fullmatch(r"\d+(?:\.\d+)?", tok):
                built.append(tok)
            elif re.fullmatch(r"[()+\-*/%,]", tok):
                built.append(tok)
            elif tok.startswith("[") and tok.endswith("]"):
                series = col_ref(tok)
                var = f"__v{len(local_ctx)}__"
                local_ctx[var] = series
                built.append(var)
            else:
                key = re.sub(r"[^A-Za-z0-9_]", "_", tok)
                if key in local_ctx:
                    built.append(key)
                else:
                    built.append(tok)
        expr_built = "".join(built)

        # Inyectar funciones
        for fn, impl in func_map.items():
            local_ctx[f"__{fn}__"] = (lambda f: f)(impl)
        for fn, impl in func_map_es.items():
            local_ctx[fn] = (lambda f: f)(impl)

        # Evaluación
        try:
            val = eval(expr_built, {"np": np, "pd": pd}, local_ctx)
        except Exception as e:
            return jsonify({"error": f"Error al evaluar fórmula: {e}", "expr": expr_built}), 400

        # Formateo
        def format_value(v):
            if isinstance(v, (pd.Series, np.ndarray)):
                return v
            return v

        if mode == 'row':
            # Añadir columna al DataFrame
            series = pd.Series(val)
            df[name] = series

            # Si se solicita formato porcentaje, formatear visualmente solo en HTML
            if fmt:
                kind, _, pattern = fmt.partition(":")
                pattern = pattern or "0.00"
                if kind == 'percent':
                    # Crear una copia para formato visual
                    df_view = df.copy()
                    # Formatear la columna recién agregada como 0.00%
                    def fmt_percent(x):
                        try:
                            if pd.isna(x):
                                return "0.00%"
                            fx = float(x)
                            # Regla determinista: si hay doble escalado evidente (|fx| > 1000),
                            # deshacer una sola vez el *100.
                            if abs(fx) > 1000:
                                fx = fx / 100.0
                            # Formatear sin clamps ni reescalados adicionales
                            return f"{fx:.2f}%"
                        except Exception:
                            return "0.00%"
                    # Aplicar formateo en la vista y también sobreescribir la columna base para que cualquier vista previa lo respete
                    df_view[name] = df_view[name].apply(fmt_percent)
                    df[name] = df[name].apply(fmt_percent)
                    html = df_view.to_html(classes="table table-bordered table-striped", index=False, escape=False)
                else:
                    html = df.to_html(classes="table table-bordered table-striped", index=False, escape=False)
            else:
                html = df.to_html(classes="table table-bordered table-striped", index=False, escape=False)

            # Resumen inferior vacío (se podría añadir footer)
            result = {
                "sheet": sheet,
                "columns": list(df.columns),
                "rows": len(df),
                "html": html
            }
        else:
            # total: reducir a escalar
            if isinstance(val, (pd.Series, np.ndarray)):
                total = float(np.nansum(val))
            else:
                total = float(val)
            result = {"sheet": sheet, "total": total}

        # Formato porcentaje opcional si detecta % en la expresión o format indica percent
        if fmt:
            kind, _, pattern = fmt.partition(":")
            pattern = pattern or "0.00"
            if mode == 'total' and 'total' in result:
                if kind == 'percent':
                    fx = float(result['total']) if result['total'] is not None else 0.0
                    # Regla determinista: si hay doble escalado evidente (|fx| > 1000), deshacer una sola vez el *100.
                    if abs(fx) > 1000:
                        fx = fx / 100.0
                    result['total_formatted'] = f"{fx:.2f}%"
                elif kind == 'number':
                    result['total_formatted'] = f"{result['total']:.2f}"
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/columns", methods=["GET"])
def columns():
    try:
        sheet = request.args.get('sheet')
        if not _LAST_BOOK:
            return jsonify({"error": "Primero sube un archivo Excel."}), 400
        if not sheet or sheet not in _LAST_BOOK:
            return jsonify({"error": "Hoja no encontrada."}), 400
        cols = list(_LAST_BOOK[sheet].columns)
        return jsonify({"sheet": sheet, "columns": cols})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/compute", methods=["POST"])
def compute():
    try:
        data = request.get_json(silent=True) or {}
        sheet = data.get("sheet")
        formulas = data.get("formulas", [])  # [{"name": "NuevaCol", "expr": "Kilos * 2"}, ...]
        group_by = data.get("group_by")  # ["Exportadora", ...]
        aggregates = data.get("aggregates")  # {"Kilos": "sum", ...}

        if not _LAST_BOOK:
            return jsonify({"error": "Primero sube un archivo Excel."}), 400
        if not sheet or sheet not in _LAST_BOOK:
            return jsonify({"error": "Hoja no encontrada. Especifica una hoja válida."}), 400

        df = _LAST_BOOK[sheet].copy()

        # Aplicar fórmulas de columnas nuevas
        allowed_funcs = {
            'abs': abs,
            'round': round,
        }
        if formulas:
            for item in formulas:
                name = item.get("name")
                expr = item.get("expr")
                if not name or not expr:
                    continue
                try:
                    # pandas.eval para expresiones vectorizadas
                    df[name] = pd.eval(expr, engine='python', parser='pandas', local_dict={**allowed_funcs, **df.to_dict(orient='series')})
                except Exception as e:
                    return jsonify({"error": f"Error en fórmula '{name}': {e}"}), 400

        # Agregaciones por grupo si se solicita
        if group_by and aggregates:
            try:
                result = df.groupby(group_by).agg(aggregates).reset_index()
            except Exception as e:
                return jsonify({"error": f"Error en agregación: {e}"}), 400
        else:
            result = df

        html = result.to_html(classes="table table-bordered table-striped", index=False, escape=False)
        return jsonify({
            "sheet": sheet,
            "columns": list(result.columns),
            "rows": len(result),
            "html": html
        })
    except Exception as e:
        print(f"💥 ERROR /compute: {e}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/upload", methods=["POST"])
def upload_file():
    try:
        # Verificar que se envió un archivo
        if 'file' not in request.files:
            return jsonify({"error": "No se envió ningún archivo"}), 400
        
        file = request.files["file"]
        
        # Verificar que se seleccionó un archivo
        if file.filename == "":
            return jsonify({"error": "No se seleccionó archivo"}), 400

        # Verificar extensión
        allowed_ext = (".xlsx", ".xls", ".xlsm", ".xlsb")
        if not file.filename.lower().endswith(allowed_ext):
            return jsonify({"error": "Solo se permiten archivos Excel (.xlsx, .xls, .xlsm, .xlsb)"}), 400

        print(f"📁 Procesando archivo: {file.filename}")
        
        # Leer Excel directamente desde la memoria
        file_bytes = io.BytesIO(file.read())
        
        print("🔍 Leyendo archivo Excel...")

        engine = _choose_engine(file.filename)
        print(f"   ↳ Engine seleccionado: {engine}")

        # Estrategia: usar la PRIMERA FILA como encabezado exactamente como en Excel, expandiendo merges si aplica.
        html_sheets = {}
        fname = file.filename.lower()
        
        # Limpiar libro previo
        global _LAST_BOOK
        _LAST_BOOK = {}

        if engine == "openpyxl":
            try:
                file_bytes.seek(0)
                wb = load_workbook(filename=file_bytes, data_only=True)
                print(f"✅ Excel leído con openpyxl. Hojas: {wb.sheetnames}")

                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]

                    # Construir mapa de celdas fusionadas
                    merged_ranges = list(ws.merged_cells.ranges)

                    def expand_merged_value(row, col):
                        for mr in merged_ranges:
                            if (row, col) in mr.cells:
                                tl_row, tl_col = mr.min_row, mr.min_col
                                return ws.cell(tl_row, tl_col).value
                        return ws.cell(row, col).value

                    max_col = ws.max_column

                    # Encabezado de dos niveles (fila 1 y fila 2), expandiendo merges
                    top_header = []
                    sub_header = []
                    for c in range(1, max_col + 1):
                        v1 = expand_merged_value(1, c)
                        v2 = expand_merged_value(2, c)
                        top_header.append("" if v1 is None else str(v1).strip())
                        sub_header.append("" if v2 is None else str(v2).strip())

                    # Combinar nombres de columnas segun reglas:
                    # - si sub_header tiene valor => usar "top - sub" si top existe, si no usar solo sub
                    # - si sub_header está vacío => usar top
                    combined_cols = []
                    for t, s in zip(top_header, sub_header):
                        if s and t:
                            combined_cols.append(f"{t} - {s}")
                        elif s and not t:
                            combined_cols.append(s)
                        else:
                            combined_cols.append(t)

                    # Limpiar 'Unnamed' y separadores residuales
                    combined_cols = _clean_columns(combined_cols)

                    # Datos desde la fila 3
                    data = []
                    for r in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=max_col, values_only=True):
                        data.append(["" if val is None else val for val in r])

                    df = pd.DataFrame(data, columns=combined_cols)
                    df = df.fillna("")
                    print(f"📊 Procesando hoja: {sheet_name} - Forma: {df.shape}")
                    print(f"   Columnas detectadas (2 filas combinadas): {list(df.columns)}")

                    # Guardar en memoria
                    _LAST_BOOK[sheet_name] = df.copy()

                    html_sheets[sheet_name] = df.to_html(
                        classes="table table-bordered table-striped",
                        index=False,
                        escape=False
                    )
            except Exception as e:
                print(f"❌ Error leyendo con openpyxl personalizado (2 niveles): {e}")
                return jsonify({"error": f"No se pudo leer el archivo .xlsx: {e}"}), 400
        else:
            # Para .xls, .xlsb: usar header=[0,1] y combinar niveles con la misma regla
            try:
                file_bytes.seek(0)
                excel_data = pd.read_excel(file_bytes, sheet_name=None, engine=engine, header=[0,1])
                print(f"✅ Excel leído con {engine}. Hojas: {list(excel_data.keys())}")
                for sheet_name, df in excel_data.items():
                    # df.columns es un MultiIndex (top, sub)
                    new_cols = []
                    for t, s in df.columns:
                        t = "" if pd.isna(t) else str(t).strip()
                        s = "" if pd.isna(s) else str(s).strip()
                        if s and t:
                            new_cols.append(f"{t} - {s}")
                        elif s and not t:
                            new_cols.append(s)
                        else:
                            new_cols.append(t)
                    df.columns = _clean_columns(new_cols)
                    df = df.fillna("")
                    print(f"📊 Procesando hoja: {sheet_name} - Forma: {df.shape}")
                    print(f"   Columnas detectadas (2 filas combinadas): {list(df.columns)}")

                    # Guardar en memoria
                    _LAST_BOOK[sheet_name] = df.copy()

                    html_sheets[sheet_name] = df.to_html(
                        classes="table table-bordered table-striped",
                        index=False,
                        escape=False
                    )
            except Exception as e:
                print(f"❌ Error leyendo con {engine}: {e}")
                return jsonify({"error": f"No se pudo leer el archivo con el motor '{engine}': {e}"}), 400

        print("🎉 Procesamiento completado exitosamente")
        return jsonify(html_sheets)

        print("🎉 Procesamiento completado exitosamente")
        return jsonify(html_sheets)

    except Exception as e:
        print(f"💥 ERROR: {str(e)}")
        print("TRACEBACK:")
        traceback.print_exc()
        
        return jsonify({"error": f"Error al procesar el archivo: {str(e)}"}), 500

@app.route("/formula/export", methods=["POST"])
def formula_export():
    try:
        data = request.get_json(force=True)
        sheet = data.get("sheet")
        expr = data.get("expr", "").strip()
        name = data.get("name", "Resultado")
        mode = data.get("mode", "row")
        fmt = data.get("format")

        if not _LAST_BOOK:
            return jsonify({"error": "Primero sube un archivo Excel."}), 400
        if not sheet or sheet not in _LAST_BOOK:
            return jsonify({"error": "Hoja no encontrada."}), 400
        if not expr:
            return jsonify({"error": "Expresión vacía."}), 400

        # Reutilizamos la evaluación de /formula replicando la lógica esencial
        import re as _re
        df = _LAST_BOOK[sheet].copy()

        def _to_series_list(args):
            out = []
            for a in args:
                if isinstance(a, pd.Series):
                    out.append(pd.to_numeric(a, errors='coerce'))
                elif isinstance(a, (list, tuple, np.ndarray)):
                    out.append(pd.to_numeric(pd.Series(a), errors='coerce'))
                else:
                    try:
                        val = pd.to_numeric(a, errors='coerce')
                    except Exception:
                        val = np.nan
                    out.append(pd.Series([val] * len(df)))
            return out

        def _row_df(args):
            return pd.concat(_to_series_list(args), axis=1)

        mode_local = mode
        func_map = {
            'SUM': lambda *args: (_row_df(args).sum(axis=1, skipna=True) if mode_local=='row' else float(np.nansum([pd.to_numeric(a, errors='coerce').sum(skipna=True) if isinstance(a, pd.Series) else pd.to_numeric(a, errors='coerce') for a in args]))),
            'AVERAGE': lambda *args: (_row_df(args).mean(axis=1, skipna=True) if mode_local=='row' else (
                (sum([pd.to_numeric(a, errors='coerce').sum(skipna=True) if isinstance(a, pd.Series) else (pd.to_numeric(a, errors='coerce') if pd.notna(pd.to_numeric(a, errors='coerce')) else 0) for a in args])) /
                max(1, sum([ (pd.to_numeric(a, errors='coerce').notna().sum() if isinstance(a, pd.Series) else (1 if pd.notna(pd.to_numeric(a, errors='coerce')) else 0)) for a in args ]))
            )),
            'MIN': lambda *args: (_row_df(args).min(axis=1, skipna=True) if mode_local=='row' else (
                np.nanmin(pd.concat(_to_series_list(args), axis=0).to_numpy())
            )),
            'MAX': lambda *args: (_row_df(args).max(axis=1, skipna=True) if mode_local=='row' else (
                np.nanmax(pd.concat(_to_series_list(args), axis=0).to_numpy())
            )),
            'COUNT': lambda *args: (_row_df(args).notna().sum(axis=1) if mode_local=='row' else sum([
                (pd.to_numeric(a, errors='coerce').notna().sum() if isinstance(a, pd.Series) else (1 if pd.notna(pd.to_numeric(a, errors='coerce')) else 0)) for a in args
            ])),
        }

        local_ctx = {}
        for col in df.columns:
            var = _re.sub(r"[^A-Za-z0-9_]", "_", col)
            base = var
            k = 1
            while var in local_ctx:
                var = f"{base}_{k}"
                k += 1
            local_ctx[var] = pd.to_numeric(df[col], errors='coerce')

        def col_ref(token: str):
            t = token.strip()
            if t in local_ctx:
                return local_ctx[t]
            t2 = t.strip("[]'\"")
            key = _re.sub(r"[^A-Za-z0-9_]", "_", t2)
            series = local_ctx.get(key)
            if series is None:
                series = pd.Series([np.nan] * len(df))
            return series

        expr_py = expr.replace(";", ",")
        for fn in func_map.keys():
            expr_py = _re.sub(rf"\b{fn}\s*\(", f"__{fn}__(", expr_py, flags=_re.IGNORECASE)
        tokens = _re.findall(r"[A-Za-z_][A-Za-z0-9_]*|\d+(?:\.\d+)?|[()+\-*/%,]|\[.*?\]", expr_py)
        built = []
        for tok in tokens:
            up = tok.upper()
            if up in {f"__{k}__" for k in func_map.keys()}:
                built.append(tok.upper())
            elif _re.fullmatch(r"\d+(?:\.\d+)?", tok):
                built.append(tok)
            elif _re.fullmatch(r"[()+\-*/%,]", tok):
                built.append(tok)
            elif tok.startswith("[") and tok.endswith("]"):
                series = col_ref(tok)
                var = f"__v{len(local_ctx)}__"
                local_ctx[var] = series
                built.append(var)
            else:
                key = _re.sub(r"[^A-Za-z0-9_]", "_", tok)
                if key in local_ctx:
                    built.append(key)
                else:
                    built.append(tok)
        expr_built = "".join(built)
        for fn, impl in func_map.items():
            local_ctx[f"__{fn}__"] = (lambda f: f)(impl)

        try:
            val = eval(expr_built, {"np": np, "pd": pd}, local_ctx)
        except Exception as e:
            return jsonify({"error": f"Error al evaluar fórmula: {e}", "expr": expr_built}), 400

        # Aplicar resultado y formateo similar a /formula
        if mode == 'row':
            series = pd.Series(val)
            df[name] = series
            # Formato percent consistente con /formula
            if fmt:
                kind, _, pattern = fmt.partition(":")
                if kind == 'percent':
                    def fmt_percent(x):
                        try:
                            if pd.isna(x):
                                return "0.00%"
                            fx = float(x)
                            if abs(fx) > 1000:
                                fx = fx / 100.0
                            return f"{fx:.2f}%"
                        except Exception:
                            return "0.00%"
                    df[name] = df[name].apply(fmt_percent)
        else:
            if isinstance(val, (pd.Series, np.ndarray)):
                total = float(np.nansum(val))
            else:
                total = float(val)
            # Para exportar en modo total, creamos un pequeño DF
            if fmt:
                kind, _, _ = fmt.partition(":")
            else:
                kind = None
            out_rows = []
            if kind == 'percent':
                fx = total
                if abs(fx) > 1000:
                    fx = fx / 100.0
                out_rows.append({"total": total, "total_formatted": f"{fx:.2f}%"})
            else:
                out_rows.append({"total": total})
            df = pd.DataFrame(out_rows)

        # Exportar a Excel en memoria
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Usamos el nombre de hoja original
            sheetname = str(sheet)[:31] or "Sheet1"
            df.to_excel(writer, index=False, sheet_name=sheetname)
        output.seek(0)

        filename = f"{sheet}-{name}-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    print("🚀 Servidor Flask iniciado en http://127.0.0.1:5000")
    print("📊 Listo para recibir archivos Excel...")
    app.run(debug=True)